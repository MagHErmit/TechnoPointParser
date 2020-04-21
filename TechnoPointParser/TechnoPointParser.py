########-----------PARSE CONSTANTS----------########
url = 'https://technopoint.ru/catalog/recipe/e351231ca6161134/2020-goda/'
count = 10 # amount of objects(smartphones)
url_api = 'https://technopoint.ru/ajax-state/price/'

########-----------EXCEL CONSTANTS----------########
topic = ['Наименование', 'Код товара', 'Цена', 'Ссылка на изображение']
title_page = 'Смартфоны 2020'
file_name = 'Смартфоны.xlsx'

########-----------GET DATA----------########
import requests
from bs4 import BeautifulSoup
import json

s = requests.Session()
s.headers.update({ # Adding user-agent for robot-block
    'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.116 Safari/537.36 OPR/67.0.3575.130'
    })
response = s.get(url)


soup = BeautifulSoup(response.text, features="lxml")
list_products = soup.find_all('div', {'class' : 'catalog-item'}) # Get basis for parse 

names = []
numbers = []
links = []
price = []
data = {"data":"{\"type\":\"price\",\"containers\":["} # Base of argument for getting prices

for i in range (count): # Parse
    names.append(list_products[i].find('div', {'class': 'product-info__title-link'}).find('a').text)
    numbers.append(list_products[i].find('span', {'data-product-param' : 'code'}).text)
    links.append(list_products[i].find('img').get('data-src'))
    data['data'] += "{\"id\":\"not-empty-%i\",\"data\":{\"product\":\"%s\"}}," % (i, list_products[i].get('data-guid')) # Forming argument for API that get prices
                                                                                                                        # "id"s can be random, but different values 
data['data'] = data['data'][:-1]
data['data'] += "]}" # End forming
    
results = [names, numbers, price, links]

token = soup.find('meta', {'name' : 'csrf-token'}).get('content') # Parse token for request to API for prices
s.headers.update({ # Adding header parametrs for getting access to API
    'x-csrf-token' : token,
    'origin' : 'https://technopoint.ru',
    'referer' : url+'no-referrer',
    'sec-fetch-mode' : 'cors',
    'content-type' : 'application/x-www-form-urlencoded',
    'x-requested-with' : 'XMLHttpRequest',
    })

ans = json.loads(s.post(url_api, data = data).text) # Get JSON with prices

for price in ans['data']['states']: # Parse prices
    results[2].append(price['data']['current'])

########-----------WRITE TO EXCEL TABLE-----------########
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

fill = PatternFill(fill_type='solid',start_color='FFD700') 
                                                                                 
font = Font(name='Calibri',bold=True,)

wb = Workbook()
ws = wb.active
ws.title = title_page
ws.append(topic)

for cellObj in ws['A1:D1']:
    for cell in cellObj:
        cell.fill = fill
        cell.font = font
for i in range(count): # Forming rows for appending
    row = []
    for j in range(0,4):
        row.append(results[j][i])
    ws.append(row)

dims = {} # Setting width depending on length of values
for row in ws.rows:
    for cell in row:
        if cell.value:
            dims[cell.column_letter] = max((dims.get(cell.column, 0), len(cell.value)))
for col, value in dims.items():
    ws.column_dimensions[col].width = value + 1

wb.save(file_name)